import pandas as pd
import csv
from io import StringIO
from typing import List, Dict, Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string


def process_sheet(sheet: Worksheet, df: pd.DataFrame) -> List[List[str]]:
    # DataFrameが空かどうかチェック
    if df.empty:
        return []

    # 印刷範囲を計算する
    print_area: Optional[str] = sheet.print_area

    if not print_area:
        # 印刷範囲が定義されていない場合、シート全体を使用
        min_col, min_row = 1, 1
        max_col, max_row = sheet.max_column, sheet.max_row
    else:
        # 印刷範囲が定義されている場合、その範囲を解析
        cell_range = print_area.split("!")[-1]  # シート名が含まれている場合に対応
        top_left, bottom_right = cell_range.split(":")

        # 印刷範囲の左上と右下のセルの座標を取得
        min_col, min_row = top_left.replace("$", "")[0], top_left.replace("$", "")[1:]
        max_col, max_row = (
            bottom_right.replace("$", "")[0],
            bottom_right.replace("$", "")[1:],
        )

        # 列文字を数値インデックスに変換
        min_col = column_index_from_string(min_col)
        min_row = int(min_row)
        max_col = column_index_from_string(max_col)
        max_row = int(max_row)

    # 非表示の列を特定
    hidden_columns: set[int] = set()
    for col_index in range(min_col, max_col + 1):
        letter = get_column_letter(col_index)
        col = sheet.column_dimensions[letter]
        if col.hidden:
            hidden_columns.add(col_index)

    # 非表示の行を特定
    hidden_rows: set[int] = set()
    for row_index in range(min_row, max_row + 1):
        row = sheet.row_dimensions[row_index]
        if row.hidden:
            hidden_rows.add(row_index)

    # 印刷範囲内のセルをループして値を取得し、出力
    sheet_data: List[List[str]] = []
    for row_index in range(min_row, max_row + 1):
        if row_index in hidden_rows:
            continue  # 非表示の行はスキップ

        row_values: List[str] = []
        for col_index in range(min_col, max_col + 1):
            if col_index in hidden_columns:
                continue  # 非表示の列はスキップ

            # DataFrameのインデックスが範囲内かチェック
            if 0 <= row_index - 1 < df.shape[0] and 0 <= col_index - 1 < df.shape[1]:
                value = df.iloc[row_index - 1, col_index - 1]
                row_values.append(str(value) if pd.notna(value) else "")
            else:
                row_values.append("")  # 範囲外の場合は空文字を追加

        sheet_data.append(row_values)

    return sheet_data


def process_excel_file(file_path: str) -> Dict[str, str]:
    results: Dict[str, str] = {}
    wb = load_workbook(file_path, data_only=True)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # pandasでシートを読み込む
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

        # シートの処理を実行
        sheet_data = process_sheet(sheet, df)

        # 処理結果をCSV形式の文字列として出力
        if sheet_data:
            output = StringIO()
            csv_writer = csv.writer(output)
            csv_writer.writerows(sheet_data)
            csv_string = output.getvalue()
            results[sheet_name] = csv_string
        else:
            print(f"No data processed for sheet: {sheet_name}")
            results[sheet_name] = ""

    return results
